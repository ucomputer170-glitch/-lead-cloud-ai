"""
excel_gen.py — Generate styled Excel files from leads
"""

from pathlib import Path
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import JOBS_DIR


HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALT_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
DATA_FONT = Font(name="Calibri", size=10)
DATA_ALIGN = Alignment(vertical="center", wrap_text=True)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

LEAD_HEADERS = [
    "S.No", "Business Name", "Owner Name", "Phone", "Email",
    "Website", "City", "State", "Niche", "Source"
]
LEAD_COL_WIDTHS = [6, 35, 25, 18, 30, 35, 18, 8, 15, 12]


def generate_excel(leads: list[dict], job_id: str, filename: str = None) -> str:
    """
    Generate styled Excel file from leads.

    Returns: file path of the generated Excel
    """
    if not filename:
        filename = f"leads_{job_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    filepath = JOBS_DIR / filename

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leads"

    for col_idx, header in enumerate(LEAD_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER

    for row_idx, lead in enumerate(leads, 2):
        row_data = [
            row_idx - 1,
            lead.get("business_name", ""),
            lead.get("owner_name", ""),
            lead.get("phone", ""),
            lead.get("email", ""),
            lead.get("website", ""),
            lead.get("city", ""),
            lead.get("state", ""),
            lead.get("niche", ""),
            lead.get("source", ""),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border = BORDER
            if row_idx % 2 == 0:
                cell.fill = ALT_FILL

    for i, width in enumerate(LEAD_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(LEAD_HEADERS))}{len(leads)+1}"
    ws.freeze_panes = "A2"

    wb.save(str(filepath))
    return str(filepath)
